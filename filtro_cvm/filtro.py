# BIBLIOTECAS
import pandas as pd
import zipfile
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from tkinter import messagebox, filedialog
import sys

try:
    # PARÂMETROS
    year = int(sys.argv[2]) # ANO DE ANÁLISE
    cia_code = int(sys.argv[1]) # CÓDIGO CVM DA CIA
    
    relevant_data = ['BPA', 'BPP', 'DFC_MI', 'DRE'] # AJUSTAR DE ACORDO COM OS DADOS DA CVM A SEREM LIDOS (CONSOLIDADOS)
    columns = ['DENOM_CIA', 'CD_CVM', 'ORDEM_EXERC', 'DT_FIM_EXERC', 'CD_CONTA', 'DS_CONTA', 'VL_CONTA'] # COLUNAS RELEVANTES
    
    columns_to_use = ['CD_CONTA', 
                      'DS_CONTA_1T', 
                      'VL_CONTA_1T', 
                      'DS_CONTA_2T', 
                      'VL_CONTA_2T', 
                      'DS_CONTA', 
                      'VL_CONTA', 
                      'DS_CONTA_DFP', 
                      'VL_CONTA_DFP'] # LISTA DE COLUNAS QUE SERÃO USADAS
    new_columns = {
        'CD_CONTA': 'CÓDIGO DA CONTA',
        'DS_CONTA_1T': 'DESCRIÇÃO DA CONTA 1T',
        'VL_CONTA_1T': 'VALOR DA CONTA 1T (R$ Mil)',
        'DS_CONTA_2T': 'DESCRIÇÃO DA CONTA 2T',
        'VL_CONTA_2T': 'VALOR DA CONTA 2T (R$ Mil)',
        'DS_CONTA': 'DESCRIÇÃO DA CONTA 3T',
        'VL_CONTA': 'VALOR DA CONTA 3T (R$ Mil)',
        'DS_CONTA_DFP': 'DESCRIÇÃO DA CONTA DFP',
        'VL_CONTA_DFP': 'VALOR DA CONTA DFP (R$ Mil)'} # NOVOS NOMES DAS COLUNAS
    
    periods = [f'{year}-03-31', f'{year}-06-30', f'{year}-09-30', f'{year}-12-31'] # PERÍODOS (1T, 2T, 3T E DFP)
    
    # DFP
    path = os.path.join(os.getcwd(), 'dados_cvm_dfp') # PATH PARA A PASTA DE DADOS
    
    dataframes_dfp = {} # DICIONÁRIO DE DATAFRAME PARA SALVAR OS DADOS RELEVANTES
    
    with zipfile.ZipFile(path + f'\\dfp_cia_aberta_{year}.zip', 'r') as zipf: # ITERANDO PELOS DADOS E SALVANDO NO DICIONÁRIO DE DATAFRAMES
            for relevant in relevant_data:
                with zipf.open(f'dfp_cia_aberta_{relevant}_con_{year}.csv', 'r') as csv_file:
                    print(csv_file)
                    temp_df_dfp = pd.read_csv(csv_file, 
                                     sep=';', 
                                     encoding='ISO-8859-1',
                                     usecols=columns
                                    )
                    temp_df_dfp = temp_df_dfp[temp_df_dfp['CD_CVM'] == cia_code]
                    temp_df_dfp = temp_df_dfp[temp_df_dfp['ORDEM_EXERC'] != 'PENÚLTIMO']
                    dataframes_dfp[relevant] = temp_df_dfp
    
    cia_name = temp_df_dfp['DENOM_CIA'].iloc[0] # NOME DA CIA
    
    # ITR
    path = os.path.join(os.getcwd(), 'dados_cvm_itr') # PATH PARA A PASTA DE DADOS
    
    dataframes_itr = {} # DICIONÁRIO DE DATAFRAME PARA SALVAR OS DADOS RELEVANTES
    
    with zipfile.ZipFile(path + f'\\itr_cia_aberta_{year}.zip', 'r') as zipf: # ITERANDO PELOS DADOS E SALVANDO NO DICIONÁRIO DE DATAFRAMES
            for relevant in relevant_data:
                with zipf.open(f'itr_cia_aberta_{relevant}_con_{year}.csv', 'r') as csv_file:
                    print(csv_file)
                    temp_df_itr = pd.read_csv(csv_file, 
                                     sep=';', 
                                     encoding='ISO-8859-1',
                                     usecols=columns
                                    )
                    temp_df_itr = temp_df_itr[temp_df_itr['CD_CVM'] == cia_code]
                    temp_df_itr = temp_df_itr[temp_df_itr['ORDEM_EXERC'] != 'PENÚLTIMO']
                    dataframes_itr[relevant] = temp_df_itr
    
    # FUNÇÃO PARA FILTRAR OS DADOS
    def create_excel():
    
        # BALANÇO PATRIMONIAL ATIVO
        df_bpa_1t = dataframes_itr['BPA'][dataframes_itr['BPA']['DT_FIM_EXERC'] == periods[0]]
        df_bpa_2t = dataframes_itr['BPA'][dataframes_itr['BPA']['DT_FIM_EXERC'] == periods[1]]
        df_bpa_3t = dataframes_itr['BPA'][dataframes_itr['BPA']['DT_FIM_EXERC'] == periods[2]]
        df_bpa_dfp = dataframes_dfp['BPA']
    
        df_merged_bpa = pd.merge(df_bpa_1t, df_bpa_2t[columns[3:]], how='outer', on='CD_CONTA', suffixes=('_1T', '_2T'))
        df_merged_bpa = pd.merge(df_merged_bpa, df_bpa_3t[columns[3:]], how='outer', on='CD_CONTA', suffixes=('', '_3T'))
        df_merged_bpa = pd.merge(df_merged_bpa, df_bpa_dfp[columns[3:]], how='outer', on='CD_CONTA', suffixes=('', '_DFP'))
    
        # BALANÇO PATRIMONIAL PASSIVO
        df_bpp_1t = dataframes_itr['BPP'][dataframes_itr['BPP']['DT_FIM_EXERC'] == periods[0]]
        df_bpp_2t = dataframes_itr['BPP'][dataframes_itr['BPP']['DT_FIM_EXERC'] == periods[1]]
        df_bpp_3t = dataframes_itr['BPP'][dataframes_itr['BPP']['DT_FIM_EXERC'] == periods[2]]
        df_bpp_dfp = dataframes_dfp['BPP']
    
        df_merged_bpp = pd.merge(df_bpp_1t, df_bpp_2t[columns[3:]], how='outer', on='CD_CONTA', suffixes=('_1T', '_2T'))
        df_merged_bpp = pd.merge(df_merged_bpp, df_bpp_3t[columns[3:]], how='outer', on='CD_CONTA', suffixes=('', '_3T'))
        df_merged_bpp = pd.merge(df_merged_bpp, df_bpp_dfp[columns[3:]], how='outer', on='CD_CONTA', suffixes=('', '_DFP'))
    
        # DEMONSTRAÇÃO DO FLUXO DE CAIXA
        df_dfc_1t = dataframes_itr['DFC_MI'][dataframes_itr['DFC_MI']['DT_FIM_EXERC'] == periods[0]]
        df_dfc_2t = dataframes_itr['DFC_MI'][dataframes_itr['DFC_MI']['DT_FIM_EXERC'] == periods[1]]
        df_dfc_3t = dataframes_itr['DFC_MI'][dataframes_itr['DFC_MI']['DT_FIM_EXERC'] == periods[2]]
        df_dfc_dfp = dataframes_dfp['DFC_MI']
    
        df_merged_dfc = pd.merge(df_dfc_1t, df_dfc_2t[columns[3:]], how='outer', on='CD_CONTA', suffixes=('_1T', '_2T'))
        df_merged_dfc = pd.merge(df_merged_dfc, df_dfc_3t[columns[3:]], how='outer', on='CD_CONTA', suffixes=('', '_3T'))
        df_merged_dfc = pd.merge(df_merged_dfc, df_dfc_dfp[columns[3:]], how='outer', on='CD_CONTA', suffixes=('', '_DFP'))
    
        # DEMONSTRAÇÃO DE RESULTADO
        df_dre_1t = dataframes_itr['DRE'][dataframes_itr['DRE']['DT_FIM_EXERC'] == periods[0]]
        df_dre_2t = dataframes_itr['DRE'][dataframes_itr['DRE']['DT_FIM_EXERC'] == periods[1]][1::2]
        df_dre_3t = dataframes_itr['DRE'][dataframes_itr['DRE']['DT_FIM_EXERC'] == periods[2]][1::2]
        df_dre_dfp = dataframes_dfp['DRE']
    
        df_merged_dre = pd.merge(df_dre_1t, df_dre_2t[columns[3:]], how='outer', on='CD_CONTA', suffixes=('_1T', '_2T'))
        df_merged_dre = pd.merge(df_merged_dre, df_dre_3t[columns[3:]], how='outer', on='CD_CONTA', suffixes=('', '_3T'))
        df_merged_dre = pd.merge(df_merged_dre, df_dre_dfp[columns[3:]], how='outer', on='CD_CONTA', suffixes=('', '_DFP'))
    
        return df_merged_bpa, df_merged_bpp, df_merged_dfc, df_merged_dre
    
    # CUSTOMIZAÇÃO DOS DATAFRAMES
    BPA, BPP, DFC, DRE = create_excel()
    
    BPA = BPA[columns_to_use]
    BPP = BPP[columns_to_use]
    DFC = DFC[columns_to_use]
    DRE = DRE[columns_to_use]
    
    BPA.rename(columns=new_columns, inplace=True)
    BPP.rename(columns=new_columns, inplace=True)
    DFC.rename(columns=new_columns, inplace=True)
    DRE.rename(columns=new_columns, inplace=True)
    
    # GERAR O EXCEL
    excels_path = os.path.join(os.getcwd(), 'Arquivos Excel') # DIRETÓRIO PARA SALVAR O ARQUIVO

    if not os.path.exists(excels_path): # CHECANDO SE A PASTA JÁ EXISTE
        os.makedirs(excels_path)

    os.chdir(excels_path)
    
    with pd.ExcelWriter(f'{cia_name}_{year}.xlsx', engine='openpyxl') as writer:
        BPA.to_excel(writer, sheet_name='BPA', index=False)
        BPP.to_excel(writer, sheet_name='BPP', index=False)
        DFC.to_excel(writer, sheet_name='DFC', index=False)
        DRE.to_excel(writer, sheet_name='DRE', index=False)
    
    # CUSTOMIZAÇÃO DO EXCEL
    wb = load_workbook(f'{cia_name}_{year}.xlsx')
    
    color_mappings = {
        'A': 'DA9694',
        'B': 'C4D79B',
        'C': 'C4D79B',
        'D': 'B1A0C7',
        'E': 'B1A0C7',
        'F': '92CDDC',
        'G': '92CDDC',
        'H': 'FABF8F',
        'I': 'FABF8F'} # COLORMAP
    
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')) # ESTILO DE BORDAS
    
    for sheet_name in wb.sheetnames: # PINTANDO AS COLUNAS (CÓDIGO DA CONTA: #DA9694, 1T: #C4D79B, 2T: #B1A0C7, 3T: #92CDDC, DFP: #FABF8F)
        sheet = wb[sheet_name] 
        for column, color in color_mappings.items():
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            for cell in sheet[column]:
                cell.fill = fill
                
        for col in range(1, 10): # COLOCANDO GRIDS
            column_letter = sheet.cell(row=1, column=col).column_letter
            for cell in sheet[column_letter]:
                cell.border = border
    
    wb.save(f'{cia_name}_{year}.xlsx')

    messagebox.showinfo('Arquivo salvo', f'Arquivo salvo em:\n{excels_path}')

except Exception as e:
    messagebox.showinfo('ERRO', 'Operação inválida\n\n- Verifique o código CVM\n\n- Verifique o ano\n\n- A empresa pode não ter dados disponíveis para o ano selecionado', icon='error')